#!/usr/bin/env python3
"""
Script para exportar pedido de compra para Excel usando o novo template de grades vinculadas aos itens.
Preenche as grades nas linhas 14-21 e vincula aos itens nas colunas X, AA, AD, AG, AJ.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import sys
import json
import os

# ─── Mapeamento de células ───────────────────────────────────────────────────
CELL_MAP = {
    'numero_pedido': 'N2',
    'role': 'AA2',
    'data': 'AN2',
    'marca': 'N3',
    'representante': 'AA3',
    'telefone': 'AN3',
    'fornecedor': 'N4',
    'email': 'AA4',
    'fat_inicio': 'AA5',
    'fat_fim': 'AH5',
    'prazo_1': 'N5',
    'prazo_2': 'Q5',
    'prazo_3': 'T5',
    'venc_1': 'N6',
    'venc_2': 'Q6',
    'venc_3': 'T6',
    'desconto': 'Z6',
    'markup': 'AF6',
}

GRADE_ROWS = { 'A': 14, 'B': 15, 'C': 16, 'D': 17, 'E': 18, 'F': 19, 'G': 20, 'H': 21 }
ITEM_START_ROW = 36
ITEM_COLS = {
    'referencia': 'C',
    'tipo': 'H',
    'cor1': 'R',
    'cor2': 'S',
    'cor3': 'T',
    'modelo': 'U',
    'custo': 'AL',
    'venda': 'AO',
}

PEDIDO_COL_MAP = {0: 'X', 1: 'AA', 2: 'AD', 3: 'AG', 4: 'AJ'}

# ─── Funções auxiliares ───────────────────────────────────────────────────────
def format_venc_display(fat_fim, prazo_dias):
    if not fat_fim: return ''
    try:
        dt = datetime.strptime(fat_fim, '%Y-%m-%d')
        venc = dt + timedelta(days=prazo_dias)
        mes = venc.strftime('%b/%y').lower()
        return f"{venc.strftime('%d/%m/%Y')} - {prazo_dias} dias {mes}"
    except: return ''

def get_column_for_size(size, categoria):
    """
    Retorna a coluna Excel para um tamanho específico.
    FEM (33-42): AB-AK
    MASC (37-48): AB-AM
    INF (16-32): AB-AR
    """
    try:
        val = int(size)
    except:
        return None
    
    # Coluna AB é o índice 28 (A=1, B=2, ..., Z=26, AA=27, AB=28)
    START_COL = 28
    
    if categoria == 'FEM':
        tamanhos = list(range(33, 43))
        if val in tamanhos:
            idx = tamanhos.index(val)
            return get_column_letter(START_COL + idx)
    elif categoria == 'MASC':
        tamanhos = list(range(37, 49))
        if val in tamanhos:
            idx = tamanhos.index(val)
            return get_column_letter(START_COL + idx)
    elif categoria == 'INF':
        tamanhos = list(range(16, 33))
        if val in tamanhos:
            idx = tamanhos.index(val)
            return get_column_letter(START_COL + idx)
    return None

# ─── Função principal ─────────────────────────────────────────────────────────
def export_to_excel(template_path, output_path, order_data, items_data, sub_orders_data):
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
    else:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
    
    # Cabeçalho principal
    ws[CELL_MAP['numero_pedido']] = order_data.get('numero_pedido') or ''
    ws[CELL_MAP['role']] = 'Comprador' if order_data.get('user_role') == 'comprador' else 'Gerente'
    ws[CELL_MAP['data']] = datetime.now().strftime('%d/%m/%Y')
    ws[CELL_MAP['marca']] = order_data.get('marca') or ''
    ws[CELL_MAP['representante']] = order_data.get('representante') or ''
    ws[CELL_MAP['telefone']] = order_data.get('telefone') or ''
    ws[CELL_MAP['fornecedor']] = order_data.get('fornecedor') or ''
    ws[CELL_MAP['email']] = order_data.get('email') or ''
    ws[CELL_MAP['fat_inicio']] = order_data.get('fat_inicio') or ''
    ws[CELL_MAP['fat_fim']] = order_data.get('fat_fim') or ''
    ws[CELL_MAP['desconto']] = order_data.get('desconto', 0) / 100.0  # Assumindo porcentagem
    ws[CELL_MAP['markup']] = order_data.get('markup', 0)
    
    prazos = order_data.get('prazos', [])
    fat_fim = order_data.get('fat_fim')
    prazo_cells = ['prazo_1', 'prazo_2', 'prazo_3']
    venc_cells = ['venc_1', 'venc_2', 'venc_3']
    for i, (prazo_key, venc_key) in enumerate(zip(prazo_cells, venc_cells)):
        if i < len(prazos):
            ws[CELL_MAP[prazo_key]] = prazos[i]
            ws[CELL_MAP[venc_key]] = format_venc_display(fat_fim, prazos[i])
    
    # Exportar Itens e vincular Grades
    for idx, item in enumerate(items_data):
        row = ITEM_START_ROW + idx
        ws[f"{ITEM_COLS['referencia']}{row}"] = item.get('referencia')
        ws[f"{ITEM_COLS['tipo']}{row}"] = item.get('tipo')
        ws[f"{ITEM_COLS['cor1']}{row}"] = item.get('cor1')
        ws[f"{ITEM_COLS['cor2']}{row}"] = item.get('cor2') or item.get('cor1')
        ws[f"{ITEM_COLS['cor3']}{row}"] = item.get('cor3') or item.get('cor1')
        ws[f"{ITEM_COLS['modelo']}{row}"] = item.get('modelo')
        ws[f"{ITEM_COLS['custo']}{row}"] = item.get('custo')
        ws[f"{ITEM_COLS['venda']}{row}"] = f"=AL{row}*$AF$6"
        
        # Preencher Grades do Item (Linhas 14-21)
        # Nota: Se múltiplos itens usarem as mesmas letras de grade (A, B...), 
        # o último item processado subscreverá os valores nas linhas 14-21.
        # No template atual, as grades parecem ser globais ou limpas a cada exportação.
        grades_do_item = item.get('grades', {})
        for letter, grade_data in grades_do_item.items():
            if letter in GRADE_ROWS:
                grade_row = GRADE_ROWS[letter]
                cat = grade_data.get('cat')
                qtds = grade_data.get('qtds', {})
                for size, qty in qtds.items():
                    col = get_column_for_size(size, cat)
                    if col:
                        ws[f"{col}{grade_row}"] = qty
        
        # Vincular ao Pedido (Colunas X, AA, AD, AG, AJ)
        for ped_idx, sub in enumerate(sub_orders_data):
            col = PEDIDO_COL_MAP.get(ped_idx)
            if not col: continue
            
            # Buscar grades deste item neste sub-pedido
            grades_selecionadas = []
            for icg in sub.get('itensComGrades', []):
                if icg.get('itemIdx') == idx:
                    grades_selecionadas = [g.get('letter') for g in icg.get('grades', [])]
                    break
            
            if grades_selecionadas:
                ws[f"{col}{row}"] = ",".join(grades_selecionadas)

    # Lojas por Sub-Pedido
    for ped_idx, sub in enumerate(sub_orders_data):
        if ped_idx in [0,1,2,3,4]:
            row_loja = 23 + ped_idx
            lojas = sub.get('lojas_numeros', [])
            ws[f'C{row_loja}'] = ", ".join(map(str, lojas))

    wb.save(output_path)
    print(json.dumps({'status': 'success', 'output': output_path}))

if __name__ == '__main__':
    if len(sys.argv) < 5:
        print(json.dumps({'status': 'error', 'message': 'Usage: export_buy_order.py <template> <output> <order_json> <items_json> [suborders_json]'}))
        sys.exit(1)
    
    template_path = sys.argv[1]
    output_path = sys.argv[2]
    order_json = sys.argv[3]
    items_json = sys.argv[4]
    sub_orders_json = sys.argv[5] if len(sys.argv) > 5 else '[]'
    
    try:
        order_data = json.loads(order_json)
        items_data = json.loads(items_json)
        sub_orders_data = json.loads(sub_orders_json)
        export_to_excel(template_path, output_path, order_data, items_data, sub_orders_data)
    except Exception as e:
        print(json.dumps({'status': 'error', 'message': str(e)}))
        sys.exit(1)
